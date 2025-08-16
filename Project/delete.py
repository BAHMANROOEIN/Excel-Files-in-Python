import openpyxl , os

def delete () :
    os.system("cls")
    load = openpyxl.load_workbook("contacts.xlsx")
    sheet_DB = load["DB"]

    for cell in sheet_DB :
        print(str(cell[0].value) + " | " + cell[1].value + " | " + str(cell[2].value ))
        print("===============================")
    if sheet_DB.max_row == 1:
         print("No Contacts...!")
         print("===============================")
    else:
        inp_name = input("Please Enter Name : ")
        list_ID = []
        os.system("cls")
        for cell_2 in sheet_DB :
            name = cell_2[1].value
            if inp_name == name:
                    list_ID.append(cell_2[0].value)
                    print(str(cell_2[0].value) + " | " + cell_2[1].value + " | " + str(cell_2[2].value ))
                    print("===============================")
                    
        if len(list_ID) == 0 :
            os.system("cls")
            print("No Contact...")
        else:
            inp_ID = int(input("Please Enter Delete ID : "))
            os.system("cls")
            if(inp_ID in list_ID) == False :
                print("Error ID ... ")
            else:
                os.system("cls")
                remaining = sheet_DB.max_row - inp_ID
                sheet_DB.delete_rows(inp_ID)
                i = 1
                ro = sheet_DB.max_row
                while i <= remaining :
                    old = sheet_DB.cell(row=ro , column=1).value
                    sheet_DB.cell(row=ro , column=1).value = old - 1
                    ro-=1                
                    i+=1

                load.save("contacts.xlsx")
                print("Ok.....")
                os.system("start contacts.xlsx")

