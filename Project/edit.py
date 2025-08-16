import openpyxl , os

def edit () :
    os.system("cls")
    load = openpyxl.load_workbook("contacts.xlsx")
    sheet_DB = load["DB"]

    for cell in sheet_DB :
        print(str(cell[0].value) + " | " + cell[1].value + " | " + str(cell[2].value ))
        print("===============================")
    
    inp_name = input("Please Enter Name : ")
    list_ID = []
    os.system("cls")
    for cell_2 in sheet_DB :
        name = cell_2[1].value
        if inp_name == name:
                list_ID.append(cell_2[0].value)
                print(str(cell_2[0].value) + " | " + cell_2[1].value + " | " + str(cell_2[2].value ))
                print("===============================")
                
    if len(list_ID) == 0 :
         os.system("cls")
         print("No Contact...")
    else:
        inp_ID = int(input("Please Enter ID : "))
        os.system("cls")
        if(inp_ID in list_ID) == False :
            print("Error ID ... ")
        else:
            os.system("cls")
            inp_name_new = input("Please Enter New Name : ")
            sheet_DB.cell(row=inp_ID , column=2).value = inp_name_new
            inp_number_new = input("Please Enter New Phone Number : ")
            sheet_DB.cell(row=inp_ID , column=3).value = inp_number_new
            os.system("cls")
            load.save("contacts.xlsx")
            print("Submit")
            os.system("start contacts.xlsx")
            
            