import openpyxl , os , create

def c_contact () :
    os.system("cls")
    create.create_ex("contacts" , "DB")
    load = openpyxl.load_workbook("contacts.xlsx")
    sheet_DB = load["DB"]
    inp_name = input("Plase Enter Name : ")
    inp_number = input("Plase Enter Phone Number : ")
    max_row = sheet_DB.max_row + 1
    sheet_DB.cell(row=max_row , column=1 ).value = max_row
    sheet_DB.cell(row=max_row , column=2 ).value = inp_name
    sheet_DB.cell(row=max_row , column=3 ).value = inp_number
    load.save("contacts.xlsx")
    os.system("cls")
    print("Submit")
    os.system("start contacts.xlsx")

